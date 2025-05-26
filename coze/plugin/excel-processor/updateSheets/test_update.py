import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
from io import BytesIO
import ast
from datetime import datetime
import tos
import os
import tempfile

def findUpdateValue(cell_value, rows):
    for row in rows:
        if len(row) >= 5 and row[0].strip() == cell_value:
            return row[2], row[3], row[4] 
    return None, None, None

def handler(content, endpoint, region, bucket_name, base_path, ak, sk):
    update_data = ast.literal_eval(content)
    rows = update_data['data']

    url = "https://pub-kylin.tos-cn-beijing.volces.com/9836/template/margin.xlsx"
    response = requests.get(url)
    wb = openpyxl.load_workbook(BytesIO(response.content))
    sheet = wb["Sheet1"] 
    for row in sheet.iter_rows(min_row=4, max_col=1):
        cell_value = row[0].value.strip() if row[0].value else ""
        update_value = findUpdateValue(cell_value, rows)
        if update_value[0] is not None and update_value[1] is not None and update_value[2] is not None:
            sheet.cell(row=row[0].row, column=3, value=update_value[0]) 
            sheet.cell(row=row[0].row, column=4, value=update_value[0]) 
            sheet.cell(row=row[0].row, column=5, value=update_value[0]) 

    filename = f"margin-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    object_key = os.path.join(base_path, filename)
    tos_client = tos.TosClientV2(ak, sk, endpoint, region)
    saved_file = ""    

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_filename = temp_file.name
        wb.save(temp_filename)
        try:
            result = tos_client.put_object_from_file(bucket_name, object_key, temp_filename)
            saved_file = f"https://{bucket_name}.{endpoint}/{object_key}"
        except Exception as e:
            print(f"Error saving file: {e}")


    print(saved_file)
content = "{'header': [[' 利 润 表 '], [' 会计月：2023-12'], [' 币种：人民币元 ']], 'data': [[' 项目 ', ' 行次 ', ' 本期金额 ', ' 本年累计金额 ', ' 上年累计金额 '], [' 一、营业收入 ', 1, 0, 0, 0], [' 减：营业成本 ', 2, 0, 0, 0], [' 税金及附加 ', 3, 448054.935294118, 5983540.75882353, 1095857.92352941], [' 销售费用 ', 4, 8285233.85882353, 14696253.8058824, 7355762.64705882], [' 管理费用 ', 5, 1233557.18235294, 2486720.30588235, 1722740.09411765], [' 财务费用 ', 6, -1577891.82941176, -5363174.33529412, -425165.517647059], [' 其中：利息费用 ', 7, 3160.47647058824, 41626.1647058824, 50012.6470588235], [' 利息收入 ', 8, -1585383.73529412, -5417358.48823529, -478545.423529412], [' 加：其他收益 ', 9, 10237.6470588235, 43568.6352941176, 22947.1705882353], [' 投资收益（损失以 “－” 填列）', 10, 0, 0, 0], [' 其中：对联营企业和合营企业的投资收益 ', 11, 0, 0, 0], [' 以摊余成本计量的金融资产终止确认收益（损失以 “－” 填列）', 12, 0, 0, 0], [' 净敞口套期收益（损失以 “－” 填列）', 13, 0, 0, 0], [' 公允价值变动收益（损失以 “－” 填列）', 14, 0, 0, 0], [' 信用减值损失（损失以 “－” 填列）', 15, 0, 0, 0], [' 资产减值损失（损失以 “－” 填列）', 16, 0, 0, 0], [' 资产处置收益（损失以 “－” 填列）', 17, 0, 0, 0], [' 二、营业利润（亏损以 “－” 填列）', 18, -8378716.5, -17759771.9, -9726247.97647059], [' 加：营业外收入 ', 19, 45294.1176470588, 240294.782352941, 81765.8588235294], [' 减：营业外支出 ', 20, 0, 0, 0], [' 三、利润总额（亏损总额以 “－” 填列）', 21, -8333422.38235294, -17519477.1176471, -9644482.11764706], [' 减：所得税费用 ', 22, -2083355.59411765, -4385083.72352941, -2394586.75882353], [' 四、净利润（净亏损以 “－” 填列）', 23, -6250066.78823529, -13134393.3941176, -7249895.35882353], [' （一）持有经营净利润（净亏损以 “－” 填列）', 24, 0, 0, 0], [' （二）终止经营净利润（净亏损以 “－” 填列）', 25, 0, 0, 0], [' 五、其他综合收益的税后净额 ', 26, 0, 0, 0], [' （一）不能重分类进损益的其他综合收益 ', 27, 0, 0, 0], [' 1. 重新计量设定受益计划变动额 ', 28, 0, 0, 0], [' 2. 权益法下不能转损益的其他综合收益 ', 29, 0, 0, 0], [' 3. 其他权益工具投资公允价值变动 ', 30, 0, 0, 0], [' 4. 企业自身信用风险风险公允价值变动 ', 31, 0, 0, 0], [' （二）将重分类进损益的其他综合收益 ', 33, 0, 0, 0], [' 1. 权益法下可转损益的其他综合收益 ', 34, 0, 0, 0], [' 2. 其他债权投资公允价值变动 ', 35, 0, 0, 0], [' 3. 金融资产重分类计入其他综合收益的金额 ', 36, 0, 0, 0], [' 4. 其他债权投资行用减值准备 ', 37, 0, 0, 0], [' 5. 现金流量套期储备 ', 38, 0, 0, 0], [' 6. 外币财务报表这算差额 ', 39, 0, 0, 0], [' 六、综合收益总额 ', 41, -6250066.78823529, -13134393.3941176, -7249895.35882353], [' 七、每股收益：', 42, 0, 0, 0], [' （一）基本每股收益 ', 43, 0, 0, 0], [' （二）稀释每股收益 ', 44, 0, 0, 0]], 'sheet': ' 利润表 ', 'file': 'https://pub-kylin.tos-cn-beijing.volces.com/93368/BCME-20231231.xlsx'}"


ak = os.getenv('TOS_ACCESS_KEY')
sk = os.getenv('TOS_SECRET_KEY')
endpoint = os.getenv('TOS_ENDPOINT')
region = os.getenv('TOS_REGION')
bucket_name = os.getenv('TOS_BUCKET')

handler(content, endpoint, region, bucket_name, "9836", ak, sk)
