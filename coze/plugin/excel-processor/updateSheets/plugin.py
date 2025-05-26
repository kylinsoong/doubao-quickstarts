from runtime import Args
from typings.ExcelUpdate.ExcelUpdate import Input, Output
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
from io import BytesIO
import ast
from datetime import datetime
import tos
import os
import tempfile

def findUpdateValue(cell_value, rows):
    for row in rows:
        if len(row) >= 5 and row[0].strip() == cell_value:
            return row[2], row[3], row[4]
    return None, None, None

"""
Each file needs to export a function named `handler`. This function is the entrance to the Tool.

Parameters:
args: parameters of the entry function.
args.input - input parameters, you can get test input value by args.input.xxx.
args.logger - logger instance used to print logs, injected by runtime.

Remember to fill in input/output in Metadata, it helps LLM to recognize and use tool.

Return:
The return data of the function, which should match the declared output parameters.
"""
def handler(args: Args[Input])->Output:

    content = args.input.content
    endpoint = args.input.endpoint
    region = args.input.region
    bucket_name = args.input.bucket_name
    base_path = args.input.base_path
    ak = args.input.ak
    sk = args.input.sk

    update_data = ast.literal_eval(content)
    rows = update_data['data']

    url = "https://pub-kylin.tos-cn-beijing.volces.com/9836/template/margin.xlsx"
    response = requests.get(url)
    wb = openpyxl.load_workbook(BytesIO(response.content))
    sheet = wb["Sheet1"]
    for row in sheet.iter_rows(min_row=4, max_col=1):
        cell_value = row[0].value.strip() if row[0].value else ""
        update_value = findUpdateValue(cell_value, rows)
        if update_value[0] is not None and update_value[1] is not None and update_value[2] is not None:
            sheet.cell(row=row[0].row, column=3, value=update_value[0])
            sheet.cell(row=row[0].row, column=4, value=update_value[0])
            sheet.cell(row=row[0].row, column=5, value=update_value[0])

    filename = f"margin-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    object_key = os.path.join(base_path, filename)
    tos_client = tos.TosClientV2(ak, sk, endpoint, region)
    saved_file = ""

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_filename = temp_file.name
        wb.save(temp_filename)
        try:
            result = tos_client.put_object_from_file(bucket_name, object_key, temp_filename)
            saved_file = f"https://{bucket_name}.{endpoint}/{object_key}"
        except Exception as e:
            print(f"Error saving file: {e}")

    return {"updated_excel": saved_file}
